import requests 
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook



mainSite = "https://www.imdb.com"
site = "https://www.imdb.com/find?q="

class Crawler():
    def WriteDetail(self):
        print("start write detail")
        filmListBook = load_workbook("films.xlsx")
        filmListActive = filmListBook.active
        filmCount = 1;

        #new book
        filmsDetail = Workbook()
        sheetDetail = filmsDetail.active
        sheetDetail.append(("Name","Director","Release Date","Starring 1","Starring 2","Starring 3",
            "Starring 4","Starring 5","Starring 6","Starring 7","Starring 8",
            "Starring 9","Starring 10",
            "Genres 1","Genres 2","Genres 3","Genres 4",
            "Rating","Gross - $"))


        for row in filmListActive.iter_rows(min_row=1, min_col=1, max_row=1852, max_col=2):
            print("**************")
            i = 1;
            print("Film Count ",filmCount)
            date="";
            name="";
            Director = "";
            grossData = "";
            Rating = "";
            StarringArray = [];
            GenresArray = [];



            for cell in row:
                if i==1:
                    name = cell.value
                    print(cell.value ,end=" ")
                    filmCount+=1;
                else:
                    date = cell.value
                    #print(cell.value ,end=" ")
                i = 0;
            print() # satir atlamasi icin hucrelerde
            #print(name,date , end=" ")


            dateArray = date.split(", ")
            year = dateArray[1]
            #print(year)
            url = site + name + " " + year
            print(url)
            url_oku = requests.get(url)
            soup = BeautifulSoup(url_oku.content, 'html.parser')

            try:
                main = soup.find('td', {'class':'result_text'})
                hrefValue = main.a.get("href")
            except:
                continue
           
            detailUrl = mainSite + hrefValue
            print(detailUrl)

            url_detail = requests.get(detailUrl)
            soupDetail = BeautifulSoup(url_detail.content, 'html.parser')

        ####################
            try:
                RatingContent = soupDetail.find('span', {'itemprop':'ratingValue'})
                Rating = RatingContent.text
                print("Rating:")
                print(Rating)
            except IOError:
                continue
            except AttributeError:
                continue
        ############
            try:
                StarringContent = soupDetail.findAll('td', {'class':''})
                Starring = StarringContent
                c=1;
                print("Starring:")
                for item in Starring:
                    if c==11:
                        break
                    staringData = item.a.text
                    StarringArray.append(staringData)
                    print(staringData)
                    c += 1;
            except IOError:
                StarringArray = ["None"]
            except AttributeError:
                StarringArray = ["None"]
        
        ###########
            try:
                GenresContent = soupDetail.findAll('div', {'class':'see-more inline canwrap'})
                Genres = GenresContent
                title = ""
                print("Genres:")
                for item in Genres:
                    data = item.findAll()
                    title = item.h4.text
                    actorCount = 1;
                    for last in data:
                        if title == "Genres:":
                            if last.text != "Genres:" and last.text != "|":
                                if actorCount == 5:
                                    break
                                actorCount += 1;
                                genresData = last.text
                                GenresArray.append(genresData)
                                print(genresData)
            except IOError:
                GenresArray = ["None"]
            except AttributeError:
                GenresArray = ["None"]
        #############
                
            #StudioContent = soupDetail.find('div', {'class':''})
            #Studio = StudioContent.a.text

        #############
            try:
                DirectorContent = soupDetail.find('div', {'class':'credit_summary_item'})
                Director = DirectorContent.a.text
                print("Director:")
                print(Director)
            except IOError:
                continue
            except AttributeError:
                continue
        ############
            try:
                GrossContent = soupDetail.findAll('div', {'class':'txt-block'})
                Gross = GrossContent
                title2 =""
                for item in Gross:
                    title2 = item.h4.text
                    if title2 == "Cumulative Worldwide Gross:":
                        print(item.text)

                        fullGross = item.text
                        grossArray = fullGross.split('$')

                        grossData = grossArray[1]
                        break
            except AttributeError:
                grossData = ""
            except IOError:
                 grossData = ""

                
        ############
          

         
            genresCount = len(GenresArray)
            if genresCount == 0:
                for item in range(4):
                    GenresArray.append(" ")
            elif  genresCount == 1:
                for item in range(3):
                    GenresArray.append(" ")
            elif  genresCount == 2:
                for item in range(2):
                    GenresArray.append(" ")
            elif  genresCount == 3:
                for item in range(1):
                    GenresArray.append(" ")
            
            starringCount = len(StarringArray)

            if starringCount == 0:
                for item in range(10):
                    StarringArray.append(" ")
            elif starringCount == 1:
                for item in range(9):
                    StarringArray.append(" ")
            elif starringCount == 2:
                for item in range(8):
                    StarringArray.append(" ")
            elif starringCount == 3:
                for item in range(7):
                    StarringArray.append(" ")
            elif starringCount == 4:
                for item in range(6):
                    StarringArray.append(" ")
            elif starringCount == 5:
                for item in range(5):
                    StarringArray.append(" ")
            elif starringCount == 6:
                for item in range(4):
                    StarringArray.append(" ")
            elif starringCount == 7:
                for item in range(3):
                    StarringArray.append(" ")
            elif starringCount == 8:
                for item in range(2):
                    StarringArray.append(" ")
            elif starringCount == 9:
                for item in range(1):
                    StarringArray.append(" ")
                    

            #print(date, "- ",name, "-",Director,"-",grossData,"-",Rating,"-",str(GenresArray),"-",str(StarringArray))


            print("_______End______")
                    

            
            

            sheetDetail.append((name,Director,date,StarringArray[0],StarringArray[1],StarringArray[2],
            StarringArray[3],StarringArray[4],StarringArray[5],StarringArray[6],StarringArray[7],
            StarringArray[8],StarringArray[9],
            GenresArray[0],GenresArray[1],GenresArray[2],GenresArray[3],
            Rating,grossData))

            filmsDetail.save("filmsDetail.xlsx")
        filmsDetail.close()


